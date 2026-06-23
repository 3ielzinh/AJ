"""
Comando: importar_monitoramento
================================
Importa as demandas das abas EVOLUTIVO, CORRETIVO e IND. DE RUBR.
de uma planilha Excel (.xlsm) para o banco de dados.

Uso:
    python manage.py importar_monitoramento
    python manage.py importar_monitoramento --arquivo /caminho/para/planilha.xlsm
    python manage.py importar_monitoramento --limpar   # apaga registros antes de importar

O comando é idempotente quando usado sem --limpar: registros com mesmo
nome + tipo não são duplicados (update_or_create pelo par nome/tipo).

Regra de negócio — cálculo de dias:
    A coluna "QTD. DIAS EM ESPERA" da planilha é ignorada.
    O valor é calculado dinamicamente pelo model (property `dias_espera`).
"""

import re
from datetime import date, datetime
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from monitoramento.models import Demanda, TipoDemanda, StatusDemanda

# Caminho padrão da planilha (Downloads do usuário atual)
PLANILHA_PADRAO = Path.home() / "Downloads" / "Monitoramento CGPJU (1).xlsm"

# Mapeamento aba → tipo de demanda
ABAS = {
    "EVOLUTIVO":    TipoDemanda.EVOLUTIVA,
    "CORRETIVO":    TipoDemanda.CORRETIVA,
    "IND. DE RUBR.": TipoDemanda.INDICACAO_RUBRICA,
    "CRIAÇÃO OBJ.": TipoDemanda.CRIACAO_OBJETO,
}

# Mapeamento para a aba CONCLUIDAS (usa coluna OBSERVAÇÕES como discriminador de tipo)
TIPOS_CONCLUIDAS = {
    "EVOLUTIVO":    TipoDemanda.EVOLUTIVA,
    "CORRETIVO":    TipoDemanda.CORRETIVA,
    "IND. DE RUBR.": TipoDemanda.INDICACAO_RUBRICA,
    "CRIAÇÃO OBJ.": TipoDemanda.CRIACAO_OBJETO,
}

# Normalização de status (planilha usa capitalização inconsistente)
STATUS_MAP = {
    "em andamento": StatusDemanda.EM_ANDAMENTO,
    "em Andamento": StatusDemanda.EM_ANDAMENTO,
    "concluído":    StatusDemanda.CONCLUIDA,
    "concluida":    StatusDemanda.CONCLUIDA,
    "não iniciado": StatusDemanda.NAO_INICIADA,
    "nao iniciado": StatusDemanda.NAO_INICIADA,
}


def _normalizar_status(raw) -> str:
    if not raw:
        return StatusDemanda.NAO_INICIADA
    key = str(raw).strip().lower()
    # Tenta match exato normalizado
    for k, v in STATUS_MAP.items():
        if k.lower() == key:
            return v
    # Fallback parcial
    if "andamento" in key:
        return StatusDemanda.EM_ANDAMENTO
    if "conclu" in key:
        return StatusDemanda.CONCLUIDA
    return StatusDemanda.NAO_INICIADA


def _to_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val.date() if isinstance(val, datetime) else val
    if isinstance(val, str):
        val = val.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                continue
    return None


def _str(val, max_len: int = 0) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    if max_len and len(s) > max_len:
        s = s[:max_len]
    return s


def _to_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def _to_bool(val) -> bool:
    if val is None:
        return False
    return str(val).strip().upper() in ("SIM", "S", "TRUE", "1", "X", "YES")


class Command(BaseCommand):
    help = "Importa demandas da planilha Monitoramento CGPJU para o banco de dados."

    def add_arguments(self, parser):
        parser.add_argument(
            "--arquivo",
            type=str,
            default=str(PLANILHA_PADRAO),
            help=f"Caminho para a planilha .xlsm (padrão: {PLANILHA_PADRAO})",
        )
        parser.add_argument(
            "--limpar",
            action="store_true",
            default=False,
            help="Apaga TODOS os registros antes de importar (reimportação limpa).",
        )

    def handle(self, *args, **options):
        try:
            import openpyxl  # noqa: PLC0415
        except ImportError as exc:
            raise CommandError(
                "openpyxl não está instalado. Execute: pip install openpyxl"
            ) from exc

        arquivo = Path(options["arquivo"])
        if not arquivo.exists():
            raise CommandError(f"Arquivo não encontrado: {arquivo}")

        self.stdout.write(f"Abrindo: {arquivo}")

        try:
            wb = openpyxl.load_workbook(
                arquivo,
                read_only=True,
                data_only=True,
                keep_vba=False,
            )
        except Exception as exc:
            raise CommandError(f"Erro ao abrir a planilha: {exc}") from exc

        if options["limpar"]:
            count, _ = Demanda.objects.all().delete()
            self.stdout.write(self.style.WARNING(f"  {count} registro(s) removido(s)."))

        total_criados = 0
        total_atualizados = 0

        with transaction.atomic():
            for nome_aba, tipo in ABAS.items():
                if nome_aba not in wb.sheetnames:
                    self.stdout.write(self.style.WARNING(f"  Aba '{nome_aba}' não encontrada — ignorada."))
                    continue

                ws = wb[nome_aba]
                rows = list(ws.iter_rows(values_only=True))

                if not rows:
                    self.stdout.write(f"  Aba '{nome_aba}': vazia — ignorada.")
                    continue

                # Mapeia colunas pelo header
                header = [str(c).upper().strip() if c else "" for c in rows[0]]
                col = {h: i for i, h in enumerate(header)}

                def _get(row, campo):
                    idx = col.get(campo)
                    return row[idx] if idx is not None and idx < len(row) else None

                dados = rows[1:]
                criados = atualizados = ignorados = 0

                for row in dados:
                    # Ignora linhas completamente vazias
                    if not any(c for c in row):
                        continue

                    nome = _str(_get(row, "NOME DA DEMANDA"), max_len=2000)
                    if not nome:
                        ignorados += 1
                        continue

                    status     = _normalizar_status(_get(row, "STATUS ATUAL"))
                    data_ini   = _to_date(_get(row, "INICIO DA DEMANDA"))
                    data_fim   = _to_date(_get(row, "FIM DA DEMANDA"))
                    proc_sei   = _str(_get(row, "PROCESSO SEI"), max_len=300)
                    proc_rel   = _str(_get(row, "PROCESSO RELACIONADO"), max_len=2000)
                    alm        = _str(_get(row, "ALM"), max_len=100)
                    local      = _str(_get(row, "LOCALIZAÇÃO"), max_len=500)
                    reiterada  = _to_bool(_get(row, "DEMANDA REITERADA"))
                    dt_reiter  = _to_date(_get(row, "DATA REITERAÇÃO"))
                    prior      = _to_int(_get(row, "PRIORIZAÇÃO"))
                    obs        = _str(_get(row, "OBSERVAÇÕES"), max_len=10000)
                    classif    = _str(_get(row, "CLASSIFICAÇÃO"), max_len=200)

                    defaults = {
                        "status":             status,
                        "data_inicio":        data_ini,
                        "data_conclusao":     data_fim,
                        "classificacao":      classif,
                        "processo_sei":       proc_sei,
                        "processo_relacionado": proc_rel,
                        "alm":                alm,
                        "localizacao":        local,
                        "reiterada":          reiterada,
                        "data_reiteracao":    dt_reiter,
                        "priorizacao":        prior,
                        "observacoes":        obs,
                    }

                    _, created = Demanda.objects.update_or_create(
                        nome=nome,
                        tipo=tipo,
                        defaults=defaults,
                    )

                    if created:
                        criados += 1
                    else:
                        atualizados += 1

                total_criados    += criados
                total_atualizados += atualizados
                self.stdout.write(
                    f"  [{nome_aba}] {criados} criado(s), {atualizados} atualizado(s), {ignorados} ignorado(s)."
                )

            # ── Aba CONCLUIDAS (tipo discriminado pela coluna OBSERVAÇÕES) ──
            aba_conc = "CONCLUIDAS"
            if aba_conc in wb.sheetnames:
                ws = wb[aba_conc]
                rows = list(ws.iter_rows(values_only=True))
                if rows:
                    header = [str(c).upper().strip() if c else "" for c in rows[0]]
                    col = {h: i for i, h in enumerate(header)}

                    def _get_c(row, campo):
                        idx = col.get(campo)
                        return row[idx] if idx is not None and idx < len(row) else None

                    criados = atualizados = ignorados = 0
                    for row in rows[1:]:
                        if not any(c for c in row):
                            continue
                        nome = _str(_get_c(row, "NOME DA DEMANDA"), max_len=2000)
                        if not nome:
                            ignorados += 1
                            continue
                        # Tipo vem da última coluna (OBSERVAÇÕES)
                        tipo_raw = _str(_get_c(row, "OBSERVAÇÕES")).upper().strip()
                        tipo = TIPOS_CONCLUIDAS.get(tipo_raw)
                        if tipo is None:
                            ignorados += 1
                            continue

                        defaults = {
                            "status":           StatusDemanda.CONCLUIDA,
                            "data_inicio":      _to_date(_get_c(row, "INICIO DA DEMANDA")),
                            "data_conclusao":   _to_date(_get_c(row, "FIM DA DEMANDA")),
                            "processo_sei":     _str(_get_c(row, "PROCESSO SEI"), max_len=300),
                            "processo_relacionado": _str(_get_c(row, "PROCESSO RELACIONADO"), max_len=2000),
                            "localizacao":      _str(_get_c(row, "LOCALIZAÇÃO"), max_len=500),
                            "reiterada":        _to_bool(_get_c(row, "DEMANDA REITERADA")),
                            "observacoes":      _str(_get_c(row, "OBSERVAÇÃO"), max_len=10000),
                            # Campos sem coluna nesta aba
                            "classificacao":    "",
                            "alm":              "",
                        }

                        _, created = Demanda.objects.update_or_create(
                            nome=nome,
                            tipo=tipo,
                            defaults=defaults,
                        )
                        if created:
                            criados += 1
                        else:
                            atualizados += 1

                    total_criados     += criados
                    total_atualizados += atualizados
                    self.stdout.write(
                        f"  [{aba_conc}] {criados} criado(s), {atualizados} atualizado(s), {ignorados} ignorado(s)."
                    )
            else:
                self.stdout.write(self.style.WARNING(f"  Aba '{aba_conc}' não encontrada — ignorada."))

        self.stdout.write(
            self.style.SUCCESS(
                f"\nImportação concluída: {total_criados} criado(s), {total_atualizados} atualizado(s)."
            )
        )
