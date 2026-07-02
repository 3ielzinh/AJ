from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
import unicodedata

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models import Count

from monitoramento.models import Demanda, ObjetoGestao, StatusDemanda, TipoDemanda


HEADER_ALIASES = {
    "id_objeto": {"ID DO OBJETO"},
    "nome_objeto": {"NOME DO OBJETO"},
    "descricao": {"DESCRICAO"},
    "grupo_objetos": {"GRUPO DE OBJETOS"},
    "ativo": {"ATIVO"},
    "data_encerramento": {"DATA ENCERRAMENTO"},
    "processo_sei": {"PROCESSO SEI"},
    "ajs_ativas": {"AJS ATIVAS"},
    "tipo_objeto": {"TIPO DE OBJETO"},
    "carater": {"CARATER"},
    "fluxo_confirmacao": {"FLUXO DE CONFIRMACAO"},
    "passivel_absorcao": {"PASSIVEL DE ABSORCAO"},
    "tema": {"TEMA"},
    "subtema": {"SUBTEMA"},
    "pedido_inicial": {"PEDIDO INICIAL"},
    "limite_maximo_objeto": {"LIMITE MAXIMO DO OBJETO"},
    "situacao": {"SITUACAO"},
    "data": {"DATA"},
    "observacao": {"OBSERVACAO"},
}

FIELDS_TO_PRESERVE_IN_OBS = [
    "ID DO OBJETO",
    "DESCRICAO",
    "ATIVO",
    "DATA ENCERRAMENTO",
    "AJS ATIVAS",
    "TIPO DE OBJETO",
    "CARATER",
    "FLUXO DE CONFIRMACAO",
    "PASSIVEL DE ABSORCAO",
    "TEMA",
    "SUBTEMA",
    "PEDIDO INICIAL",
    "LIMITE MAXIMO DO OBJETO",
    "OBSERVACAO ORIGINAL",
]

SIM_VALUES = {"SIM", "S", "TRUE", "1", "YES"}
NAO_VALUES = {"NAO", "N", "FALSE", "0", "NO"}

# Valores conhecidos; os demais sao reportados como inesperados no dry-run.
KNOWN_SITUACAO_VALUES = {
    "ANALISADO",
    "EM ANALISE",
    "PENDENTE",
    "EM ANDAMENTO",
    "CONCLUIDO",
    "CONCLUIDA",
    "NAO INICIADA",
    "ATIVO",
    "INATIVO",
}


@dataclass
class ParsedRow:
    sheet: str
    row_number: int
    nome: str
    grupo: str
    processo_sei: str
    ativo_raw: str
    situacao_raw: str
    data_raw: str
    data_encerramento_raw: str
    data_parsed: date | None
    data_encerramento_parsed: date | None
    observacoes: str
    raw_values: dict[str, str]


def _normalize_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = " ".join(text.upper().split())
    return text


def _clean_str(value, max_len: int | None = None) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if max_len and len(text) > max_len:
        return text[:max_len]
    return text


def _parse_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _build_column_map(header_row: tuple) -> dict[str, int]:
    normalized_header = [_normalize_text(cell) for cell in header_row]
    column_map: dict[str, int] = {}
    for canonical, aliases in HEADER_ALIASES.items():
        alias_norm = {_normalize_text(alias) for alias in aliases}
        for idx, normalized in enumerate(normalized_header):
            if normalized in alias_norm:
                column_map[canonical] = idx
                break
    return column_map


def _find_header_row(rows: list[tuple], max_scan: int = 40) -> tuple[int, dict[str, int]]:
    limit = min(max_scan, len(rows))
    for idx in range(limit):
        row = rows[idx]
        column_map = _build_column_map(row)
        if "nome_objeto" in column_map and len(column_map) >= 4:
            return idx, column_map
    raise ValueError("Nao foi possivel detectar cabecalho com 'NOME DO OBJETO'.")


def _cell(row: tuple, column_map: dict[str, int], key: str):
    idx = column_map.get(key)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _build_observacoes(sheet_name: str, row_number: int, raw_values: dict[str, str]) -> str:
    lines = [
        "Importacao Gestao de Objetos (planilha):",
        f"- ABA ORIGEM: {sheet_name}",
        f"- LINHA ORIGEM: {row_number}",
    ]

    for field_name in FIELDS_TO_PRESERVE_IN_OBS:
        value = raw_values.get(field_name, "")
        if value:
            lines.append(f"- {field_name}: {value}")

    return "\n".join(lines)


class Command(BaseCommand):
    help = (
        "Importa abas de Gestao de Objetos para ObjetoGestao, "
        "com dry-run analitico e importacao idempotente."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "arquivo",
            type=str,
            help="Caminho da planilha .xlsx com as abas de Gestao de Objetos.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            default=False,
            help="Apenas analisa e mostra o que seria feito, sem gravar no banco.",
        )
        parser.add_argument(
            "--status-rule",
            choices=["none", "ativo"],
            default="none",
            help=(
                "Regra para status na importacao real: "
                "mantida por compatibilidade; nao altera ObjetoGestao."
            ),
        )
        parser.add_argument(
            "--mapear-data",
            choices=["none", "inicio"],
            default="none",
            help=(
                "Mapeamento para coluna DATA na importacao real: "
                "mantido por compatibilidade; nao altera ObjetoGestao."
            ),
        )
        parser.add_argument(
            "--chave-idempotencia",
            choices=["nome-tipo", "nome-grupo-tipo"],
            default="nome-tipo",
            help=(
                "Chave usada no update_or_create: "
                "nome-tipo (usa nome) ou nome-grupo-tipo (usa nome+grupo)."
            ),
        )

    def handle(self, *args, **options):
        try:
            import openpyxl  # noqa: PLC0415
        except ImportError as exc:
            raise CommandError("Dependencia ausente: instale openpyxl>=3.1.") from exc

        arquivo = Path(options["arquivo"])
        dry_run = bool(options["dry_run"])
        status_rule = options["status_rule"]
        mapear_data = options["mapear_data"]
        chave_idempotencia = options["chave_idempotencia"]

        if not arquivo.exists():
            raise CommandError(f"Arquivo nao encontrado: {arquivo}")

        self.stdout.write(f"Abrindo planilha: {arquivo}")
        wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)

        sheet_names = [s for s in wb.sheetnames if _normalize_text(s) != "OBJETO"]
        ignored_sheet_names = [s for s in wb.sheetnames if _normalize_text(s) == "OBJETO"]
        if ignored_sheet_names:
            self.stdout.write(self.style.WARNING("Aba 'Objeto' ignorada por regra."))

        if not sheet_names:
            raise CommandError("Nenhuma aba de dados encontrada para importar.")

        parsed_rows: list[ParsedRow] = []
        sheet_summary: dict[str, dict[str, int]] = defaultdict(
            lambda: {"criadas": 0, "atualizadas": 0, "ignoradas": 0, "erros": 0}
        )
        errors_by_row: list[str] = []
        rows_without_name: list[str] = []
        invalid_dates: list[str] = []
        grupo_counter: Counter[str] = Counter()
        active_values_counter: Counter[str] = Counter()
        situacao_values_counter: Counter[str] = Counter()
        duplicate_name_groups: dict[str, set[str]] = defaultdict(set)

        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                sheet_summary[sheet_name]["ignoradas"] += 1
                continue

            try:
                header_idx, column_map = _find_header_row(rows)
            except ValueError as exc:
                errors_by_row.append(f"{sheet_name}: cabecalho nao identificado ({exc})")
                sheet_summary[sheet_name]["erros"] += 1
                continue

            data_rows = rows[header_idx + 1 :]
            for offset, row in enumerate(data_rows, start=header_idx + 2):
                try:
                    if not any(cell is not None and str(cell).strip() for cell in row):
                        sheet_summary[sheet_name]["ignoradas"] += 1
                        continue

                    nome = _clean_str(_cell(row, column_map, "nome_objeto"), max_len=2000)
                    if not nome:
                        rows_without_name.append(f"{sheet_name}: linha {offset}")
                        sheet_summary[sheet_name]["ignoradas"] += 1
                        continue

                    grupo = _clean_str(_cell(row, column_map, "grupo_objetos"), max_len=200)
                    processo_sei = _clean_str(_cell(row, column_map, "processo_sei"), max_len=300)
                    ativo_raw = _clean_str(_cell(row, column_map, "ativo"), max_len=60)
                    situacao_raw = _clean_str(_cell(row, column_map, "situacao"), max_len=100)
                    data_raw = _clean_str(_cell(row, column_map, "data"), max_len=60)
                    data_encerramento_raw = _clean_str(
                        _cell(row, column_map, "data_encerramento"),
                        max_len=60,
                    )

                    data_parsed = _parse_date(_cell(row, column_map, "data"))
                    if data_raw and data_parsed is None:
                        invalid_dates.append(f"{sheet_name}: linha {offset} (DATA='{data_raw}')")

                    data_encerramento_parsed = _parse_date(_cell(row, column_map, "data_encerramento"))
                    if data_encerramento_raw and data_encerramento_parsed is None:
                        invalid_dates.append(
                            f"{sheet_name}: linha {offset} (DATA ENCERRAMENTO='{data_encerramento_raw}')"
                        )

                    ativo_norm = _normalize_text(ativo_raw)
                    if ativo_norm:
                        active_values_counter[ativo_norm] += 1

                    situacao_norm = _normalize_text(situacao_raw)
                    if situacao_norm:
                        situacao_values_counter[situacao_norm] += 1

                    duplicate_name_groups[_normalize_text(nome)].add(grupo or "(sem grupo)")
                    grupo_counter[grupo or "(sem grupo)"] += 1

                    raw_values = {
                        "ID DO OBJETO": _clean_str(_cell(row, column_map, "id_objeto"), max_len=80),
                        "DESCRICAO": _clean_str(_cell(row, column_map, "descricao"), max_len=2000),
                        "ATIVO": ativo_raw,
                        "DATA ENCERRAMENTO": data_encerramento_raw,
                        "AJS ATIVAS": _clean_str(_cell(row, column_map, "ajs_ativas"), max_len=500),
                        "TIPO DE OBJETO": _clean_str(_cell(row, column_map, "tipo_objeto"), max_len=500),
                        "CARATER": _clean_str(_cell(row, column_map, "carater"), max_len=500),
                        "FLUXO DE CONFIRMACAO": _clean_str(
                            _cell(row, column_map, "fluxo_confirmacao"),
                            max_len=500,
                        ),
                        "PASSIVEL DE ABSORCAO": _clean_str(
                            _cell(row, column_map, "passivel_absorcao"),
                            max_len=500,
                        ),
                        "TEMA": _clean_str(_cell(row, column_map, "tema"), max_len=500),
                        "SUBTEMA": _clean_str(_cell(row, column_map, "subtema"), max_len=500),
                        "PEDIDO INICIAL": _clean_str(_cell(row, column_map, "pedido_inicial"), max_len=2000),
                        "LIMITE MAXIMO DO OBJETO": _clean_str(
                            _cell(row, column_map, "limite_maximo_objeto"),
                            max_len=500,
                        ),
                        "OBSERVACAO ORIGINAL": _clean_str(_cell(row, column_map, "observacao"), max_len=4000),
                    }

                    parsed_rows.append(
                        ParsedRow(
                            sheet=sheet_name,
                            row_number=offset,
                            nome=nome,
                            grupo=grupo,
                            processo_sei=processo_sei,
                            ativo_raw=ativo_raw,
                            situacao_raw=situacao_raw,
                            data_raw=data_raw,
                            data_encerramento_raw=data_encerramento_raw,
                            data_parsed=data_parsed,
                            data_encerramento_parsed=data_encerramento_parsed,
                            observacoes=_build_observacoes(sheet_name, offset, raw_values),
                            raw_values=raw_values,
                        )
                    )
                except Exception as exc:  # noqa: BLE001
                    sheet_summary[sheet_name]["erros"] += 1
                    errors_by_row.append(f"{sheet_name}: linha {offset} ({exc})")

        if not parsed_rows:
            self._print_preflight(
                sheet_names=sheet_names,
                total_candidates=0,
                grupo_counter=grupo_counter,
                duplicate_name_groups=duplicate_name_groups,
                rows_without_name=rows_without_name,
                invalid_dates=invalid_dates,
                active_values_counter=active_values_counter,
                situacao_values_counter=situacao_values_counter,
                errors_by_row=errors_by_row,
            )
            raise CommandError("Nenhum registro candidato encontrado para importacao.")

        nome_norm_to_name = {
            _normalize_text(row.nome): row.nome
            for row in parsed_rows
            if _normalize_text(row.nome)
        }
        candidate_names = sorted({row.nome for row in parsed_rows})
        existing_qs = ObjetoGestao.objects.filter(nome__in=candidate_names)

        if chave_idempotencia == "nome-tipo":
            db_duplicates = list(
                existing_qs.values("nome")
                .annotate(qtd=Count("id"))
                .filter(qtd__gt=1)
            )
            if db_duplicates:
                duplicate_preview = ", ".join(item["nome"] for item in db_duplicates[:10])
                raise CommandError(
                    "Ha nomes duplicados ja existentes em ObjetoGestao. "
                    "Resolva antes de usar update_or_create com chave nome+tipo. Exemplos: "
                    f"{duplicate_preview}"
                )
        else:
            db_duplicates = list(
                existing_qs.values("nome", "grupo")
                .annotate(qtd=Count("id"))
                .filter(qtd__gt=1)
            )
            if db_duplicates:
                duplicate_preview = ", ".join(
                    f"{item['nome']} ({item['grupo']})"
                    for item in db_duplicates[:10]
                )
                raise CommandError(
                    "Ha duplicidades ja existentes para a chave nome+classificacao+tipo. "
                    "Resolva antes de importar. Exemplos: "
                    f"{duplicate_preview}"
                )

        existing_by_key = {
            self._build_lookup_key_from_objeto(d, chave_idempotencia): d
            for d in existing_qs
        }

        possible_duplicate_names = {
            nome_norm_to_name.get(name_norm, name_norm): sorted(groups)
            for name_norm, groups in duplicate_name_groups.items()
            if len(groups) > 1
        }

        self._print_preflight(
            sheet_names=sheet_names,
            total_candidates=len(parsed_rows),
            grupo_counter=grupo_counter,
            duplicate_name_groups=duplicate_name_groups,
            rows_without_name=rows_without_name,
            invalid_dates=invalid_dates,
            active_values_counter=active_values_counter,
            situacao_values_counter=situacao_values_counter,
            errors_by_row=errors_by_row,
            possible_duplicate_names=possible_duplicate_names,
        )

        self._print_status_proposal(parsed_rows)
        self._print_change_preview(
            parsed_rows,
            existing_by_key,
            status_rule,
            mapear_data,
            chave_idempotencia,
        )

        if possible_duplicate_names and chave_idempotencia == "nome-tipo":
            self.stdout.write(
                self.style.ERROR(
                    "Importacao bloqueada: ha nomes repetidos em grupos diferentes na planilha. "
                    "Defina estrategia de chave (ex.: nome+grupo+tipo) antes de gravar."
                )
            )
            if dry_run:
                return
            raise CommandError("Importacao cancelada por duplicidade de nome entre grupos.")

        if dry_run:
            self.stdout.write(self.style.SUCCESS("Dry-run concluido sem gravacao."))
            return

        self.stdout.write(self.style.WARNING("Importacao real iniciada (com transaction.atomic)."))

        with transaction.atomic():
            created_total = 0
            updated_total = 0
            ignored_total = 0

            seen_keys: set[tuple[str, str]] = set()
            for row in parsed_rows:
                sheet_stats = sheet_summary[row.sheet]
                if not row.nome:
                    ignored_total += 1
                    sheet_stats["ignoradas"] += 1
                    continue

                lookup_key = self._build_lookup_key_from_row(row, chave_idempotencia)
                current_obj = existing_by_key.get(lookup_key)
                defaults = self._build_defaults(row, current_obj, status_rule, mapear_data)

                if chave_idempotencia == "nome-grupo-tipo":
                    lookup_kwargs = {
                        "nome": row.nome,
                        "grupo": row.grupo,
                    }
                else:
                    lookup_kwargs = {
                        "nome": row.nome,
                    }

                _, created = ObjetoGestao.objects.update_or_create(
                    **lookup_kwargs,
                    defaults=defaults,
                )

                if created and lookup_key not in seen_keys:
                    created_total += 1
                    sheet_stats["criadas"] += 1
                    seen_keys.add(lookup_key)
                else:
                    updated_total += 1
                    sheet_stats["atualizadas"] += 1

            self.stdout.write("\nResumo por aba:")
            for sheet_name in sheet_names:
                stats = sheet_summary[sheet_name]
                self.stdout.write(
                    f"- [{sheet_name}] criadas={stats['criadas']}, "
                    f"atualizadas={stats['atualizadas']}, ignoradas={stats['ignoradas']}, erros={stats['erros']}"
                )

            self.stdout.write(
                self.style.SUCCESS(
                    "\nImportacao concluida: "
                    f"criadas={created_total}, atualizadas={updated_total}, ignoradas={ignored_total}."
                )
            )

    def _build_defaults(
        self,
        row: ParsedRow,
        current_obj: ObjetoGestao | None,
        status_rule: str,
        mapear_data: str,
    ) -> dict:
        raw = row.raw_values
        demanda_origem = self._find_demanda_origem(row)
        return {
            "id_objeto": raw.get("ID DO OBJETO", ""),
            "descricao": raw.get("DESCRICAO", ""),
            "grupo": row.grupo,
            "ativo": self._parse_bool(row.ativo_raw),
            "data_encerramento": row.data_encerramento_parsed,
            "processo_sei": row.processo_sei,
            "ajs_ativas": raw.get("AJS ATIVAS", ""),
            "tipo_objeto": raw.get("TIPO DE OBJETO", ""),
            "carater": raw.get("CARATER", ""),
            "fluxo_confirmacao": raw.get("FLUXO DE CONFIRMACAO", ""),
            "passivel_absorcao": self._parse_bool(raw.get("PASSIVEL DE ABSORCAO", "")),
            "tema": raw.get("TEMA", ""),
            "subtema": raw.get("SUBTEMA", ""),
            "pedido_inicial": raw.get("PEDIDO INICIAL", ""),
            "limite_maximo_objeto": self._parse_decimal(raw.get("LIMITE MAXIMO DO OBJETO", "")),
            "observacao": raw.get("OBSERVACAO ORIGINAL", ""),
            "aba_origem": row.sheet,
            "linha_origem": row.row_number,
            "demanda_origem": demanda_origem,
        }

    def _parse_bool(self, value: str) -> bool | None:
        normalized = _normalize_text(value)
        if normalized in SIM_VALUES:
            return True
        if normalized in NAO_VALUES:
            return False
        return None

    def _parse_decimal(self, value: str) -> Decimal | None:
        text = (value or "").strip()
        if not text:
            return None
        text = text.replace("R$", "").replace(" ", "")
        if "," in text and "." in text:
            text = text.replace(".", "").replace(",", ".")
        elif "," in text:
            text = text.replace(",", ".")
        try:
            return Decimal(text)
        except InvalidOperation:
            return None

    def _find_demanda_origem(self, row: ParsedRow) -> Demanda | None:
        return (
            Demanda.objects.filter(tipo=TipoDemanda.GESTAO_OBJETOS, nome=row.nome)
            .order_by("id")
            .first()
        )

    def _resolve_status(self, ativo_raw: str, current_status: str | None, status_rule: str) -> str:
        ativo_norm = _normalize_text(ativo_raw)

        if status_rule == "ativo":
            if ativo_norm in SIM_VALUES:
                return StatusDemanda.EM_ANDAMENTO
            if ativo_norm in NAO_VALUES:
                return StatusDemanda.CONCLUIDA

        if current_status:
            return current_status
        return StatusDemanda.NAO_INICIADA

    def _build_lookup_key_from_row(self, row: ParsedRow, chave_idempotencia: str) -> tuple[str, str]:
        if chave_idempotencia == "nome-grupo-tipo":
            return row.nome, row.grupo
        return row.nome, ""

    def _build_lookup_key_from_objeto(self, objeto: ObjetoGestao, chave_idempotencia: str) -> tuple[str, str]:
        if chave_idempotencia == "nome-grupo-tipo":
            return objeto.nome, objeto.grupo
        return objeto.nome, ""

    def _print_preflight(
        self,
        *,
        sheet_names: list[str],
        total_candidates: int,
        grupo_counter: Counter[str],
        duplicate_name_groups: dict[str, set[str]],
        rows_without_name: list[str],
        invalid_dates: list[str],
        active_values_counter: Counter[str],
        situacao_values_counter: Counter[str],
        errors_by_row: list[str],
        possible_duplicate_names: dict[str, list[str]] | None = None,
    ):
        self.stdout.write("\n=== Preflight / Dry-run ===")
        self.stdout.write(f"- Total de abas lidas: {len(sheet_names)}")
        self.stdout.write(f"- Total de registros candidatos: {total_candidates}")

        self.stdout.write("\nTotal por grupo:")
        if not grupo_counter:
            self.stdout.write("- (nenhum grupo identificado)")
        else:
            for grupo, qtd in grupo_counter.most_common():
                self.stdout.write(f"- {grupo}: {qtd}")

        self.stdout.write("\nPossiveis duplicidades por NOME DO OBJETO:")
        if possible_duplicate_names:
            for nome, groups in list(possible_duplicate_names.items())[:30]:
                self.stdout.write(f"- {nome}: grupos={', '.join(groups)}")
            if len(possible_duplicate_names) > 30:
                self.stdout.write(f"- ... e mais {len(possible_duplicate_names) - 30}")
        else:
            repeated_names = [
                name_norm
                for name_norm, groups in duplicate_name_groups.items()
                if len(groups) == 1
            ]
            if repeated_names:
                self.stdout.write(f"- Nomes repetidos no mesmo grupo: {len(repeated_names)}")
            else:
                self.stdout.write("- Nenhuma duplicidade relevante encontrada.")

        self.stdout.write(f"\nLinhas sem nome: {len(rows_without_name)}")
        for item in rows_without_name[:20]:
            self.stdout.write(f"- {item}")
        if len(rows_without_name) > 20:
            self.stdout.write(f"- ... e mais {len(rows_without_name) - 20}")

        self.stdout.write(f"\nDatas invalidas: {len(invalid_dates)}")
        for item in invalid_dates[:20]:
            self.stdout.write(f"- {item}")
        if len(invalid_dates) > 20:
            self.stdout.write(f"- ... e mais {len(invalid_dates) - 20}")

        unexpected_ativo = {
            value: qtd
            for value, qtd in active_values_counter.items()
            if value not in SIM_VALUES and value not in NAO_VALUES
        }
        self.stdout.write("\nValores inesperados em ATIVO:")
        if unexpected_ativo:
            for value, qtd in sorted(unexpected_ativo.items()):
                self.stdout.write(f"- {value}: {qtd}")
        else:
            self.stdout.write("- Nenhum.")

        unexpected_situacao = {
            value: qtd
            for value, qtd in situacao_values_counter.items()
            if value not in KNOWN_SITUACAO_VALUES
        }
        self.stdout.write("\nValores inesperados em SITUACAO:")
        if unexpected_situacao:
            for value, qtd in sorted(unexpected_situacao.items()):
                self.stdout.write(f"- {value}: {qtd}")
        else:
            self.stdout.write("- Nenhum.")

        self.stdout.write(f"\nErros de leitura por linha/aba: {len(errors_by_row)}")
        for item in errors_by_row[:20]:
            self.stdout.write(f"- {item}")
        if len(errors_by_row) > 20:
            self.stdout.write(f"- ... e mais {len(errors_by_row) - 20}")

    def _print_status_proposal(self, rows: list[ParsedRow]):
        ativo_counter = Counter(_normalize_text(row.ativo_raw) for row in rows if row.ativo_raw)
        situacao_counter = Counter(_normalize_text(row.situacao_raw) for row in rows if row.situacao_raw)
        with_data_enc = sum(1 for row in rows if row.data_encerramento_parsed)

        self.stdout.write("\n=== Proposta de regra (a validar) ===")
        self.stdout.write("- ATIVO=SIM sera gravado como ativo=True")
        self.stdout.write("- ATIVO=NAO sera gravado como ativo=False")
        self.stdout.write("- DATA ENCERRAMENTO valida sera gravada em ObjetoGestao.data_encerramento")
        self.stdout.write("- SITUACAO sera reportada no dry-run, mas nao altera status porque catalogo nao e fila")
        self.stdout.write("\nDistribuicao observada:")
        for value, qtd in ativo_counter.most_common():
            self.stdout.write(f"- ATIVO {value}: {qtd}")
        for value, qtd in situacao_counter.most_common():
            self.stdout.write(f"- SITUACAO {value}: {qtd}")
        self.stdout.write(f"- Linhas com DATA ENCERRAMENTO valida: {with_data_enc}")

    def _print_change_preview(
        self,
        rows: list[ParsedRow],
        existing_by_key: dict[tuple[str, str], ObjetoGestao],
        status_rule: str,
        mapear_data: str,
        chave_idempotencia: str,
    ):
        overwrite_counts = Counter()
        sample_changes: list[str] = []
        predicted_created = 0
        predicted_updated = 0
        seen_keys: set[tuple[str, str]] = set()

        for row in rows:
            lookup_key = self._build_lookup_key_from_row(row, chave_idempotencia)
            current_obj = existing_by_key.get(lookup_key)
            if current_obj is None and lookup_key not in seen_keys:
                predicted_created += 1
                seen_keys.add(lookup_key)
            else:
                predicted_updated += 1

            defaults = self._build_defaults(row, current_obj, status_rule, mapear_data)
            if not current_obj:
                continue

            for field in (
                "id_objeto",
                "descricao",
                "grupo",
                "ativo",
                "data_encerramento",
                "processo_sei",
                "ajs_ativas",
                "tipo_objeto",
                "carater",
                "fluxo_confirmacao",
                "passivel_absorcao",
                "tema",
                "subtema",
                "pedido_inicial",
                "limite_maximo_objeto",
                "observacao",
            ):
                old_value = getattr(current_obj, field)
                new_value = defaults[field]
                if old_value != new_value:
                    overwrite_counts[field] += 1
                    if len(sample_changes) < 25:
                        sample_changes.append(
                            f"- {row.sheet}:L{row.row_number} | {row.nome} | {field}"
                        )

        self.stdout.write("\n=== Previsao de escrita ===")
        self.stdout.write(f"- Registros previstos para criar: {predicted_created}")
        self.stdout.write(f"- Registros previstos para atualizar: {predicted_updated}")

        self.stdout.write("\nCampos com potencial sobrescrita:")
        if overwrite_counts:
            for field, qty in overwrite_counts.most_common():
                self.stdout.write(f"- {field}: {qty}")
        else:
            self.stdout.write("- Nenhuma sobrescrita detectada.")

        if sample_changes:
            self.stdout.write("\nAmostra de alteracoes:")
            for item in sample_changes:
                self.stdout.write(item)
